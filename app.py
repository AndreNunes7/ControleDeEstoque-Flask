from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}
estoque = []

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def processar_estoque(filename):
    wb = load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], filename))
    sheet = wb.active
    
    dados_estoque = []
    empty_line_count = 0  
    
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if index < 2:  
            continue
        
        if all(cell is None for cell in row):
            empty_line_count += 1
            if empty_line_count == 1:
                dados_estoque.append({'produto': '', 'quantidade': '', 'contagem': ''})
            continue
        
        if row[0] is not None and row[1] is not None:
            produto = row[0]
            quantidade = row[1]
            contagem = row[2] if len(row) > 2 else None
            dados_estoque.append({'produto': produto, 'quantidade': quantidade, 'contagem': contagem})
            empty_line_count = 0  
    
    return dados_estoque

@app.route('/', methods=['GET', 'POST'])
def home():
    global estoque
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Nenhum arquivo enviado', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('Nenhum arquivo selecionado', 'error')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            
            estoque = processar_estoque(filename)
            
            flash('Arquivo enviado com sucesso', 'success')
            return redirect(url_for('home'))
    
    return render_template('home.html', estoque=estoque)

@app.route('/editar/<produto>', methods=['POST'])
def editar_produto(produto):
    global estoque
    nova_quantidade = int(request.form['quantidade'])
    nova_contagem = int(request.form('contagem'))

    for item in estoque:
        if item['produto'] == produto:
            item['quantidade'] = nova_quantidade
            if nova_contagem:  
                item['contagem'] = nova_contagem
            break
    
    return 'OK', 200

@app.route('/excluir/<produto>')
def excluir_produto(produto):
    global estoque
    estoque = [item for item in estoque if item['produto'] != produto]
    flash(f'Produto "{produto}" excluído do estoque', 'success')
    return redirect(url_for('home'))

@app.route('/download')
def download_estoque():
    global estoque
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Estoque'
    
    sheet['A1'] = 'Produto'
    sheet['B1'] = 'Quantidade'
    sheet['C1'] = 'Contagem'

    for i, item in enumerate(estoque, start=2):
        sheet[f'A{i}'] = item['produto']
        sheet[f'B{i}'] = item['quantidade']
        sheet[f'C{i}'] = item['contagem']
    
    filename = 'estoque_atualizado.xlsx'
    wb.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
    
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
